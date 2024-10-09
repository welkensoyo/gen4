import API.njson as j
import csv
import openpyxl, xlrd
from lxml import html
import os
from io import BytesIO, StringIO
from tabula import read_pdf, convert_into
from tabula.errors import CSVParseError
import bottle
import logging

logger = logging.getLogger('sdb_app')

class PDF(object):
    def __init__(self, file=None, filename=''):
        self.file = file
        self.filename = filename
        if self.file:
            self.filename = file.filename.lower()
            if os.path.exists(self.filename):
                os.remove(self.filename)
            file.save(self.filename)
            self.result = None

    def read(self, output_format='json', multiple_tables=True, pages='all'):
        try:
            return read_pdf(self.filename, output_format=output_format, pages=pages, multiple_tables=multiple_tables, silent=True)
        except CSVParseError as e:
            # Attempt to set `names` to manage different column counts.
            column_names = [f'column_{i}' for i in range(20)]  # Assume a maximum of 20 columns
            return read_pdf(self.filename, output_format=output_format, pages=pages, multiple_tables=multiple_tables, silent=True, pandas_options={'names': column_names})


class Excel():
    def __init__(self, ext=None):
        self.file = None
        self.filename = None
        self.wb = None
        self.ext = ext
        self.delimiter = ''

    def open(self, file, data_only=False, filename=''):
        if filename:
            self.filename = os.path.basename(filename).replace(' ','_')
        if isinstance(file, bottle.FileUpload):
            self.filename = file.filename
            file = file.file
            file.seek(0)
            binary = True
        elif isinstance(file, bytes):
            binary = True
        else:
            binary = False
            self.filename = os.path.basename(file).replace(' ','_')
        self.file = file
        if not self.ext:
            self.ext = os.path.splitext(self.filename)[1].lower()
        if self.ext == '.csv':
            self._csv(binary=binary)
        if self.ext == '.xls':
            self._xls(binary=binary)
        if self.ext == '.xlsx':
            self._xlsx(binary=binary, data_only=data_only)
        if self.ext == '.pdf':
            self.wb = PDF(filename=filename).read()
        return self

    def _xlsx(self, binary=False, data_only=False):
        if binary:
            try:
                self.wb = openpyxl.load_workbook(filename=BytesIO(self.file.read()), data_only=data_only)
            except:
                self.wb = openpyxl.load_workbook(filename=BytesIO(self.file), data_only=data_only)
        if not self.wb:
            self.wb = openpyxl.load_workbook(filename=self.file, data_only=data_only)
        if self.wb:
            self.ext = '.xlsx'

    def _xls(self, binary=False):
        if binary:
            try:
                self.wb = xlrd.open_workbook(file_contents=self.file)
            except:
                self.wb = False
        if not self.wb:
            try:
                self.wb = xlrd.open_workbook(self.file)
            except:
                self.wb = False
        if not self.wb:
            try:
                self.wb = open(self.file, 'r').read()
                self.ext = '.html'
            except:
                self.delimiter = '\t'
                return self._csv(binary=binary)
        elif self.wb:
            self.ext = '.xls'
        else:
            self.ext = '.xlsx'

    def _csv(self, binary=False):
        if binary:
            try:
                try:
                    self.wb = StringIO(self.file.decode())
                except:
                    self.wb = StringIO(self.file)
            except (UnicodeDecodeError, TypeError, AttributeError):
                self.ext = '.xls'
                return self
        if self.wb:
            self.ext = '.csv'



    def ws(self, sheetname=False, header=True, check_field=None, delimiter=','):
        if not self.delimiter:
            self.delimiter = delimiter
        if self.ext == '.html':
            tree = html.fromstring(self.wb)
            tables = tree.xpath('//table')
            ws = []
            longest = 0
            index = 0
            for i, table in enumerate(tables):
                rows = table.xpath('.//tr')
                data = [row for row in [[td.text_content() for td in row.xpath('.//th')] for row in rows] if len(row) > 1]
                data.extend([row for row in [[td.text_content() for td in row.xpath('.//td')] for row in rows] if len(row) > 1])
                ws.append(data)
                x = len(data)
                if x > longest:
                    longest = x
                    index = i
            return ws[index]
        if self.ext == '.pdf':
            return self.wb
        if self.ext == '.csv':
            if self.wb:
                return self.csv_read_wb(header=header, delimiter=self.delimiter)
            return self.csv_reader(header=header, delimiter=self.delimiter)
        if self.ext == '.xls':
            if sheetname:
                ws = self.wb.get_sheet_by_name(sheetname)
            else:
                ws = self.wb.sheet_by_index(0)
            return [ws.row_values(row) for row in range(ws.nrows)]

        if not sheetname:
            ws = self.wb[self.wb.sheetnames[0]]
        else:
            try:
                ws = self.wb[sheetname]
            except KeyError: #in case the sheet doesn't exist try the first one
                ws = self.wb[self.wb.sheetnames[0]]
        xlist = []
        xa = xlist.append
        for row in ws.rows:
            xcell = []
            for cell in row:
                xcell.append(cell.value)
            if check_field:
                if xcell[check_field]:
                    xa(xcell)
            elif set(xcell) != set([None]):
                xa(xcell)
        if header == False:
            return xlist[1:]
        return xlist

    def ws_hyperlinks(self, sheetname=False, header=True):
        if not sheetname:
            ws = self.wb[self.wb.sheetnames[0]]
        else:
            ws = self.wb[sheetname]
        xlist = []
        xa = xlist.append
        for row in ws.rows:
            xcell = []
            for cell in row:
                try:
                    xcell.append(cell.hyperlink.target)
                except:
                    xcell.append(cell.value)
            xa(xcell)
        if header == False:
            return xlist[1:]
        return xlist

    def csv_gen(self, header=True):
        with open(self.file, 'r') as file:
            reader = csv.reader(file)
            if not header:
                next(reader)  # this will skip the header row
            for row in reader:
                yield row

    def csv_reader(self, header=True, check_field=False, delimiter=','):
        with open(self.file, 'r') as file:
            reader = csv.reader(file, delimiter=delimiter)
            if header == False:
                next(reader)  # this will skip the header row
            if check_field:
                return [row for row in reader if row[check_field]]
            else:
                return [row for row in reader]

    def csv_read_wb(self, header=True, check_field=False, delimiter=','):
        reader = csv.reader(self.wb, delimiter=delimiter)
        if header == False:
            next(reader)  # this will skip the header row
        if check_field:
            return [row for row in reader if row[check_field]]
        else:
            return [row for row in reader]

    def header_map(self, header):
        x = {}
        for i, h in enumerate(header):
            if h:
                x[h.lower().strip()] = i
        return x


def measures(x, filename):
    x = j.loads(x)
    measures = []
    for i in x['model']['tables']:
        if i.get('name') == 'Calcs':
            measures = i['measures']
    rows = []
    cols = ["name", "expression", "formatString", "lineageTag", "annotations"]
    rows.append(cols)
    rows.extend([[m.get(k) for k in cols] for m in measures ])
    with open(f'{filename}.csv', 'w',newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)

def get_filename(path):
    return os.path.basename(path)